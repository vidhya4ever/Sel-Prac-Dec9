# from openpyxl import Workbook
#
# wb = Workbook()
# ws = wb.active
# ws.title = 'data'
# ws['A1'] = "Name"
# ws['B1'] = "Place"
# ws['C1'] = "phonenumber"
# ws['D1'] = "Email_id"
# # print(list(ws.values))
# print(list(ws.values))
# data_list = [['nithya','chennai','9987876545','nith@gmail.com'],
#         ['vidhya','coimbatore','9987656789','vv@gmail.com'],
#         ['lakshmi','bengaluru','9989876543','lak@gmail.com'],
#         ['ramya','mysore','9987654321','rr@gmail.com']]
# for ele in data_list:
#     ws.append(ele)
# print(list(ws.values))
# wb.save('studentdata.xlsx')

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'data'
ws['A1'] = "Name"
ws['B1'] = "Place"
ws['C1'] = "phonenumber"
ws['D1'] = "Email_id"
# ws['A1'].font = Font(name='Arial', bold=True, italic=True)


print(list(ws.values))
# print(list(ws.values))
# data_list = [['nithya','chennai','9987876545','nith@gmail.com'],
#         ['vidhya','coimbatore','9987656789','vv@gmail.com'],
#         ['lakshmi','bengaluru','9989876543','lak@gmail.com'],
#         ['ramya','mysore','9987654321','rr@gmail.com']]
# for ele in data_list:
#     ws.append(ele)
# print(list(ws.values))
# wb.save('studentdata.xlsx')